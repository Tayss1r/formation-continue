"""
API endpoints for the Invitation Workflow.

- Company invitation validation & info
- Employee invitation creation (by company)
- Employee invitation validation & registration
- Coordinator: get employees for a company application
- Result file generation
- Publish results as news
"""

import os
import io
import math
from datetime import datetime, timezone
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, status, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, EmailStr
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from sqlalchemy.orm import selectinload

from ..db.database import get_session
from ..db.models import (
    User,
    UserRole,
    Company,
    CompanyApplication,
    CompanyInvitation,
    EmployeeInvitation,
    EmployeeProfile,
    EmployeeSubmission,
    EmployeeSubmissionStatus,
    ApplicationStatus,
    CallForApplicants,
    News,
)
from ..dependencies import get_current_user, RoleChecker
from ..services.invitation_service import InvitationService
from ..services.user_service import UserService
from ..utils import hash as hash_password
from ..core.config import settings

invitations_router = APIRouter()

require_company = RoleChecker([UserRole.COMPANY])
require_coordinator = RoleChecker([UserRole.COORDINATOR, UserRole.ADMIN])


# =============================================================================
# SCHEMAS
# =============================================================================

class CompanyInvitationInfo(BaseModel):
    """Info returned when a company accesses their invitation link."""
    application_id: int
    call_title: str
    call_reference: str
    company_name: str
    proposed_employee_count: int
    employees_invited: int
    token: str


class EmployeeInviteRequest(BaseModel):
    """Request to invite employees."""
    employees: List[dict]  # [{"name": "...", "email": "..."}]


class EmployeeInviteResponse(BaseModel):
    message: str
    invited_count: int


class EmployeeInvitationInfo(BaseModel):
    """Info returned when employee accesses their invitation link."""
    employee_name: str
    employee_email: str
    company_name: str
    call_title: str
    call_reference: str
    token: str
    is_used: bool


class EmployeeRegisterRequest(BaseModel):
    """Employee registration via invitation."""
    token: str
    fullname: str
    email: str
    password: str


class EmployeeRegisterResponse(BaseModel):
    message: str


class InvitedEmployeeInfo(BaseModel):
    """Info about an invited employee (for company dashboard)."""
    id: int
    name: str
    email: str
    is_used: bool
    created_at: Optional[str] = None


class CompanyInvitePageInfo(BaseModel):
    """Full info for the company invitation page."""
    application_id: int
    call_title: str
    call_reference: str
    company_name: str
    proposed_employee_count: int
    invited_employees: List[InvitedEmployeeInfo]
    token: str


# =============================================================================
# COMPANY INVITATION ENDPOINTS
# =============================================================================

@invitations_router.get("/company/validate")
async def validate_company_invitation(
    token: str = Query(...),
    session: AsyncSession = Depends(get_session),
):
    """
    Validate a company invitation token and return call/company info.
    Public endpoint (no auth required) - accessed via email link.
    """
    invitation = await InvitationService.get_company_invitation_by_token(token, session)

    valid, message = InvitationService.validate_company_invitation(invitation)
    if not valid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=message)

    app = invitation.application
    call = app.call
    company = app.company

    # Get existing employee invitations
    emp_invitations = await InvitationService.get_employee_invitations_for_company(
        invitation.id, session
    )

    invited_employees = [
        InvitedEmployeeInfo(
            id=inv.id,
            name=inv.employee_name,
            email=inv.employee_email,
            is_used=inv.is_used,
            created_at=inv.created_at.isoformat() if inv.created_at else None,
        )
        for inv in emp_invitations
    ]

    return CompanyInvitePageInfo(
        application_id=app.id,
        call_title=call.title,
        call_reference=call.reference_number,
        company_name=company.name,
        proposed_employee_count=app.proposed_employee_count,
        invited_employees=invited_employees,
        token=token,
    )


@invitations_router.post("/company/invite-employees", response_model=EmployeeInviteResponse)
async def invite_employees(
    request: EmployeeInviteRequest,
    token: str = Query(...),
    session: AsyncSession = Depends(get_session),
):
    """
    Company invites employees via their invitation token.
    Public endpoint - authenticated by the invitation token itself.
    """
    invitation = await InvitationService.get_company_invitation_by_token(token, session)

    valid, message = InvitationService.validate_company_invitation(invitation)
    if not valid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=message)

    if not request.employees:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Veuillez ajouter au moins un employé",
        )

    app = invitation.application
    max_allowed = app.proposed_employee_count or 0
    already_invited = len(await InvitationService.get_employee_invitations_for_company(invitation.id, session))
    requested_count = len(request.employees)

    if already_invited >= max_allowed:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Limite atteinte : {already_invited}/{max_allowed} employés déjà invités.",
        )

    if already_invited + requested_count > max_allowed:
        remaining = max_allowed - already_invited
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Vous ne pouvez inviter que {remaining} employé(s) supplémentaire(s). Limite totale: {max_allowed}.",
        )

    # Create employee invitations
    emp_invitations = await InvitationService.create_employee_invitations(
        company_invitation=invitation,
        employees=request.employees,
        session=session,
    )

    call = app.call
    company = app.company

    # Send emails to each employee
    for inv in emp_invitations:
        InvitationService.send_employee_invitation_email(
            employee_email=inv.employee_email,
            employee_name=inv.employee_name,
            company_name=company.name,
            call_title=call.title,
            invitation_token=inv.token,
        )

    await session.commit()

    return EmployeeInviteResponse(
        message=f"{len(emp_invitations)} invitation(s) envoyée(s) avec succès",
        invited_count=len(emp_invitations),
    )


# =============================================================================
# EMPLOYEE INVITATION ENDPOINTS
# =============================================================================

@invitations_router.get("/employee/validate")
async def validate_employee_invitation(
    token: str = Query(...),
    session: AsyncSession = Depends(get_session),
):
    """
    Validate an employee invitation token and return info for the registration page.
    Public endpoint.
    """
    invitation = await InvitationService.get_employee_invitation_by_token(token, session)

    if not invitation:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Invitation non trouvée",
        )

    # We show info even if expired/used (frontend displays message)
    comp_inv = invitation.company_invitation
    app = comp_inv.application
    call = app.call
    company = app.company

    return EmployeeInvitationInfo(
        employee_name=invitation.employee_name,
        employee_email=invitation.employee_email,
        company_name=company.name,
        call_title=call.title,
        call_reference=call.reference_number,
        token=token,
        is_used=invitation.is_used,
    )


@invitations_router.post("/employee/register", response_model=EmployeeRegisterResponse)
async def register_employee_via_invitation(
    request: EmployeeRegisterRequest,
    session: AsyncSession = Depends(get_session),
):
    """
    Register an employee account via invitation token.
    Creates user + employee profile + links to company + creates submission.
    """
    invitation = await InvitationService.get_employee_invitation_by_token(
        request.token, session
    )

    valid, message = InvitationService.validate_employee_invitation(invitation)
    if not valid:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=message)

    # Check if email already exists
    existing_user = await UserService.get_user_by_email(request.email, session)
    # Get company from invitation chain
    comp_inv = invitation.company_invitation
    app = comp_inv.application
    company = app.company

    if existing_user:
        # If account already exists, allow linking only for employee accounts.
        if existing_user.role != UserRole.EMPLOYEE:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Un compte existe déjà avec cet email mais ce n'est pas un compte employé.",
            )

        # Get or create employee profile
        profile_query = select(EmployeeProfile).where(EmployeeProfile.user_id == existing_user.id)
        profile_result = await session.execute(profile_query)
        employee_profile = profile_result.scalar_one_or_none()

        if not employee_profile:
            employee_profile = EmployeeProfile(
                user_id=existing_user.id,
                company_id=company.id,
            )
            session.add(employee_profile)
            await session.flush()
        elif not employee_profile.company_id:
            employee_profile.company_id = company.id

        # If a submission for this app/profile does not exist, create it.
        submission_query = select(EmployeeSubmission).where(
            EmployeeSubmission.company_application_id == app.id,
            EmployeeSubmission.employee_id == employee_profile.id,
        )
        submission_result = await session.execute(submission_query)
        existing_submission = submission_result.scalar_one_or_none()

        if not existing_submission:
            session.add(
                EmployeeSubmission(
                    company_application_id=app.id,
                    employee_id=employee_profile.id,
                    status=EmployeeSubmissionStatus.PENDING,
                )
            )

        invitation.is_used = True
        await session.commit()

        return EmployeeRegisterResponse(
            message="Invitation liée à votre compte existant. Connectez-vous pour déposer vos documents.",
        )

    # Generate username
    username = await UserService.generate_unique_username(request.fullname, session)

    # Create user
    new_user = User(
        username=username,
        email=request.email,
        password=hash_password(request.password),
        fullname=request.fullname,
        role=UserRole.EMPLOYEE,
        is_verified=True,  # Auto-verified via invitation
    )
    session.add(new_user)
    await session.flush()

    # Create employee profile linked to the company
    employee_profile = EmployeeProfile(
        user_id=new_user.id,
        company_id=company.id,
    )
    session.add(employee_profile)
    await session.flush()

    # Create employee submission for this application
    submission = EmployeeSubmission(
        company_application_id=app.id,
        employee_id=employee_profile.id,
        status=EmployeeSubmissionStatus.PENDING,
    )
    session.add(submission)

    # Mark invitation as used
    invitation.is_used = True

    await session.commit()

    return EmployeeRegisterResponse(
        message="Compte créé avec succès. Vous pouvez maintenant vous connecter et soumettre vos documents.",
    )


# =============================================================================
# COORDINATOR: EMPLOYEE REVIEW ENDPOINTS
# =============================================================================

@invitations_router.get("/coordinator/application/{application_id}/employees")
async def get_application_employees(
    application_id: int,
    current_user: User = Depends(require_coordinator),
    session: AsyncSession = Depends(get_session),
):
    """
    Get all employees (invited + registered + submissions) for an approved application.
    Coordinator only.
    """
    # Get the application
    app_query = (
        select(CompanyApplication)
        .options(
            selectinload(CompanyApplication.call),
            selectinload(CompanyApplication.company).selectinload(Company.user),
            selectinload(CompanyApplication.employee_submissions)
            .selectinload(EmployeeSubmission.employee)
            .selectinload(EmployeeProfile.user),
            selectinload(CompanyApplication.employee_submissions)
            .selectinload(EmployeeSubmission.documents),
        )
        .where(CompanyApplication.id == application_id)
    )
    result = await session.execute(app_query)
    app = result.scalar_one_or_none()

    if not app:
        raise HTTPException(status_code=404, detail="Candidature non trouvée")

    # Check ownership
    if current_user.role != UserRole.ADMIN and app.call.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Accès non autorisé")

    # Get invitation info
    inv_query = (
        select(CompanyInvitation)
        .options(selectinload(CompanyInvitation.employee_invitations))
        .where(CompanyInvitation.application_id == application_id)
    )
    inv_result = await session.execute(inv_query)
    company_inv = inv_result.scalar_one_or_none()

    invited_count = 0
    registered_count = 0
    employee_invitations_info = []

    if company_inv:
        for ei in company_inv.employee_invitations:
            invited_count += 1
            if ei.is_used:
                registered_count += 1
            employee_invitations_info.append({
                "id": ei.id,
                "name": ei.employee_name,
                "email": ei.employee_email,
                "is_used": ei.is_used,
                "created_at": ei.created_at.isoformat() if ei.created_at else None,
            })

    # Build submissions info
    submissions_info = []
    for sub in app.employee_submissions:
        sub_status = sub.status.value if hasattr(sub.status, "value") else sub.status
        emp_info = None
        if sub.employee and sub.employee.user:
            emp_info = {
                "id": sub.employee.id,
                "fullname": sub.employee.user.fullname,
                "email": sub.employee.user.email,
            }
        docs_info = []
        for doc in sub.documents:
            doc_status = doc.review_status.value if hasattr(doc.review_status, "value") else doc.review_status
            docs_info.append({
                "id": doc.id,
                "document_type": doc.document_type,
                "document_label": doc.document_label,
                "file_path": doc.file_path,
                "original_filename": doc.original_filename,
                "file_size": doc.file_size,
                "mime_type": doc.mime_type,
                "review_status": doc_status,
                "uploaded_at": doc.uploaded_at.isoformat() if doc.uploaded_at else None,
            })
        submissions_info.append({
            "id": sub.id,
            "status": sub_status,
            "employee": emp_info,
            "documents": docs_info,
            "created_at": sub.created_at.isoformat() if sub.created_at else None,
            "reviewed_at": sub.reviewed_at.isoformat() if sub.reviewed_at else None,
            "review_notes": sub.review_notes,
        })

    return {
        "application_id": app.id,
        "call_id": app.call.id if app.call else None,
        "call_status": app.call.status.value if app.call and hasattr(app.call.status, "value") else (app.call.status if app.call else None),
        "call_title": app.call.title if app.call else "",
        "call_reference": app.call.reference_number if app.call else "",
        "company_name": app.company.name if app.company else "",
        "proposed_employee_count": app.proposed_employee_count,
        "invited_count": invited_count,
        "registered_count": registered_count,
        "employee_invitations": employee_invitations_info,
        "submissions": submissions_info,
        "employee_required_documents": app.call.employee_required_documents if app.call else [],
    }


@invitations_router.post("/coordinator/submission/{submission_id}/approve")
async def approve_employee_submission(
    submission_id: int,
    notes: Optional[str] = None,
    current_user: User = Depends(require_coordinator),
    session: AsyncSession = Depends(get_session),
):
    """Approve an employee submission."""
    sub_query = (
        select(EmployeeSubmission)
        .options(
            selectinload(EmployeeSubmission.company_application)
            .selectinload(CompanyApplication.call),
        )
        .where(EmployeeSubmission.id == submission_id)
    )
    result = await session.execute(sub_query)
    sub = result.scalar_one_or_none()

    if not sub:
        raise HTTPException(status_code=404, detail="Soumission non trouvée")

    # Check ownership
    call = sub.company_application.call
    if current_user.role != UserRole.ADMIN and call.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Accès non autorisé")

    if sub.status in [EmployeeSubmissionStatus.APPROVED, EmployeeSubmissionStatus.REJECTED]:
        raise HTTPException(
            status_code=409,
            detail="Cette soumission a déjà une décision finale",
        )

    sub.status = EmployeeSubmissionStatus.APPROVED
    sub.reviewed_by_id = current_user.id
    sub.reviewed_at = datetime.now(timezone.utc)
    sub.review_notes = notes

    await session.commit()

    return {"message": "Soumission employé approuvée", "status": "approved"}


@invitations_router.post("/coordinator/submission/{submission_id}/reject")
async def reject_employee_submission(
    submission_id: int,
    notes: Optional[str] = None,
    current_user: User = Depends(require_coordinator),
    session: AsyncSession = Depends(get_session),
):
    """Reject an employee submission."""
    sub_query = (
        select(EmployeeSubmission)
        .options(
            selectinload(EmployeeSubmission.company_application)
            .selectinload(CompanyApplication.call),
        )
        .where(EmployeeSubmission.id == submission_id)
    )
    result = await session.execute(sub_query)
    sub = result.scalar_one_or_none()

    if not sub:
        raise HTTPException(status_code=404, detail="Soumission non trouvée")

    call = sub.company_application.call
    if current_user.role != UserRole.ADMIN and call.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Accès non autorisé")

    if sub.status in [EmployeeSubmissionStatus.APPROVED, EmployeeSubmissionStatus.REJECTED]:
        raise HTTPException(
            status_code=409,
            detail="Cette soumission a déjà une décision finale",
        )

    sub.status = EmployeeSubmissionStatus.REJECTED
    sub.reviewed_by_id = current_user.id
    sub.reviewed_at = datetime.now(timezone.utc)
    sub.review_notes = notes

    await session.commit()

    return {"message": "Soumission employé rejetée", "status": "rejected"}


# =============================================================================
# RESULT GENERATION & PUBLICATION
# =============================================================================

@invitations_router.get("/coordinator/call/{call_id}/results-data")
async def get_call_results_data(
    call_id: int,
    current_user: User = Depends(require_coordinator),
    session: AsyncSession = Depends(get_session),
):
    """
    Get summary data for result generation: approved companies + approved employees.
    """
    call_query = (
        select(CallForApplicants)
        .options(
            selectinload(CallForApplicants.applications)
            .selectinload(CompanyApplication.company),
            selectinload(CallForApplicants.applications)
            .selectinload(CompanyApplication.employee_submissions)
            .selectinload(EmployeeSubmission.employee)
            .selectinload(EmployeeProfile.user),
        )
        .where(CallForApplicants.id == call_id)
    )
    result = await session.execute(call_query)
    call = result.scalar_one_or_none()

    if not call:
        raise HTTPException(status_code=404, detail="Appel non trouvé")

    if current_user.role != UserRole.ADMIN and call.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Accès non autorisé")

    companies = []
    for app in call.applications:
        app_status = app.status.value if hasattr(app.status, "value") else app.status
        if app_status != "approved":
            continue

        employees = []
        for sub in app.employee_submissions:
            sub_status = sub.status.value if hasattr(sub.status, "value") else sub.status
            emp_name = sub.employee.user.fullname if sub.employee and sub.employee.user else "N/A"
            emp_email = sub.employee.user.email if sub.employee and sub.employee.user else "N/A"
            employees.append({
                "name": emp_name,
                "email": emp_email,
                "status": sub_status,
            })

        companies.append({
            "id": app.company.id if app.company else None,
            "name": app.company.name if app.company else "N/A",
            "industry_sector": app.company.industry_sector if app.company else "N/A",
            "proposed_employee_count": app.proposed_employee_count,
            "employees": employees,
            "approved_employees": len([e for e in employees if e["status"] == "approved"]),
        })

    return {
        "call_id": call.id,
        "call_title": call.title,
        "call_reference": call.reference_number,
        "department": call.department.value if hasattr(call.department, "value") else call.department,
        "companies": companies,
        "total_approved_companies": len(companies),
        "total_approved_employees": sum(c["approved_employees"] for c in companies),
    }


@invitations_router.get("/coordinator/call/{call_id}/generate-results")
async def generate_results_file(
    call_id: int,
    format: str = Query("excel", pattern="^(excel|pdf)$"),
    current_user: User = Depends(require_coordinator),
    session: AsyncSession = Depends(get_session),
):
    """
    Generate and download a results file (Excel or PDF) for a call.
    """
    # Reuse results-data logic
    call_query = (
        select(CallForApplicants)
        .options(
            selectinload(CallForApplicants.applications)
            .selectinload(CompanyApplication.company),
            selectinload(CallForApplicants.applications)
            .selectinload(CompanyApplication.employee_submissions)
            .selectinload(EmployeeSubmission.employee)
            .selectinload(EmployeeProfile.user),
        )
        .where(CallForApplicants.id == call_id)
    )
    result = await session.execute(call_query)
    call = result.scalar_one_or_none()

    if not call:
        raise HTTPException(status_code=404, detail="Appel non trouvé")

    if current_user.role != UserRole.ADMIN and call.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Accès non autorisé")

    rows = []
    for app in call.applications:
        app_status = app.status.value if hasattr(app.status, "value") else app.status
        if app_status != "approved":
            continue
        company_name = app.company.name if app.company else "N/A"
        for sub in app.employee_submissions:
            sub_status = sub.status.value if hasattr(sub.status, "value") else sub.status
            if sub_status != "approved":
                continue
            emp_name = sub.employee.user.fullname if sub.employee and sub.employee.user else "N/A"
            emp_email = sub.employee.user.email if sub.employee and sub.employee.user else "N/A"
            rows.append({
                "Entreprise": company_name,
                "Employé": emp_name,
                "Email": emp_email,
                "Statut Employé": sub_status,
            })

    if format == "excel":
        # Excel format
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="Module openpyxl non installé pour la génération Excel",
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Résultats"

        # Title row
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = f"Résultats – {call.title} ({call.reference_number})"
        title_cell.font = Font(size=14, bold=True)
        title_cell.alignment = Alignment(horizontal="center")

        # Date row
        ws.merge_cells("A2:D2")
        date_cell = ws["A2"]
        date_cell.value = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
        date_cell.alignment = Alignment(horizontal="center")
        date_cell.font = Font(size=10, italic=True)

        # Headers
        headers = ["Entreprise", "Employé", "Email", "Statut"]
        header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, 5):
            for col_idx, key in enumerate(["Entreprise", "Employé", "Email", "Statut Employé"], 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=row.get(key, ""))
                cell.border = thin_border

                # Color status cell
                if key == "Statut Employé":
                    if row[key] == "approved":
                        cell.fill = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")
                        cell.value = "Approuvé"
                    elif row[key] == "rejected":
                        cell.fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
                        cell.value = "Rejeté"
                    elif row[key] == "pending":
                        cell.value = "En attente"

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 4

        # Save to buffer
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=resultats_{call.reference_number}.xlsx"
            },
        )

    # PDF format
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="Module reportlab non installé pour la génération PDF",
        )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=1.5 * cm, leftMargin=1.5 * cm, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"Résultats – {call.title} ({call.reference_number})", styles["Title"]))
    story.append(Spacer(1, 0.3 * cm))
    story.append(Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", styles["Normal"]))
    story.append(Spacer(1, 0.5 * cm))

    if rows:
        data = [["Entreprise", "Employé", "Email", "Statut"]]
        for row in rows:
            data.append([
                row.get("Entreprise", ""),
                row.get("Employé", ""),
                row.get("Email", ""),
                "Approuvé",
            ])

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563EB")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("Aucun employé approuvé pour cet appel.", styles["Normal"]))

    doc.build(story)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=resultats_{call.reference_number}.pdf"
        },
    )


@invitations_router.get("/public/call/{call_id}/results-file")
async def public_results_file(
    call_id: int,
    format: str = Query("excel", pattern="^(excel|pdf)$"),
    session: AsyncSession = Depends(get_session),
):
    """
    Public download endpoint for published call results.
    Used in public news links.
    """
    call_query = (
        select(CallForApplicants)
        .options(
            selectinload(CallForApplicants.applications)
            .selectinload(CompanyApplication.company),
            selectinload(CallForApplicants.applications)
            .selectinload(CompanyApplication.employee_submissions)
            .selectinload(EmployeeSubmission.employee)
            .selectinload(EmployeeProfile.user),
        )
        .where(CallForApplicants.id == call_id)
    )
    result = await session.execute(call_query)
    call = result.scalar_one_or_none()

    if not call:
        raise HTTPException(status_code=404, detail="Appel non trouvé")

    from ..db.models import CallStatus
    call_status = call.status.value if hasattr(call.status, "value") else call.status
    if call_status != CallStatus.RESULTS_PUBLISHED.value:
        raise HTTPException(status_code=403, detail="Les résultats ne sont pas encore publiés")

    # Reuse generator logic by delegating to same file builder through direct endpoint call
    # (auth is intentionally bypassed here because call status is already validated as public).
    class _PublicUser:
        role = UserRole.ADMIN
        id = 0

    return await generate_results_file(
        call_id=call_id,
        format=format,
        current_user=_PublicUser(),
        session=session,
    )


class PublishResultsRequest(BaseModel):
    call_id: int
    title: Optional[str] = None


@invitations_router.post("/coordinator/call/{call_id}/publish-results")
async def publish_results_as_news(
    call_id: int,
    current_user: User = Depends(require_coordinator),
    session: AsyncSession = Depends(get_session),
):
    """
    Create a public news item with the results of a call.
    Includes list of admitted companies and their approved employees.
    """
    call_query = (
        select(CallForApplicants)
        .options(
            selectinload(CallForApplicants.applications)
            .selectinload(CompanyApplication.company),
            selectinload(CallForApplicants.applications)
            .selectinload(CompanyApplication.employee_submissions)
            .selectinload(EmployeeSubmission.employee)
            .selectinload(EmployeeProfile.user),
        )
        .where(CallForApplicants.id == call_id)
    )
    result = await session.execute(call_query)
    call = result.scalar_one_or_none()

    if not call:
        raise HTTPException(status_code=404, detail="Appel non trouvé")

    if current_user.role != UserRole.ADMIN and call.created_by_id != current_user.id:
        raise HTTPException(status_code=403, detail="Accès non autorisé")

    domain = (settings.DOMAIN or "localhost:8000").strip()
    if domain.startswith("http://") or domain.startswith("https://"):
        api_base = domain.rstrip("/")
    else:
        api_base = f"http://{domain.rstrip('/')}"
    pdf_link = f"{api_base}/api/v1/invitations/public/call/{call.id}/results-file?format=pdf"

    # Build news content - simple paragraph with download links only
    content = f"""<p>Les résultats de l'appel à candidatures <strong>« {call.title} »</strong> (Réf: {call.reference_number}) ont été publiés.</p>
<p>Les candidats retenus ont été notifiés par email. Vous pouvez consulter la liste complète des résultats en téléchargeant le fichier ci-dessous.</p>
<p><strong>Télécharger le fichier des résultats :</strong> <a href='{pdf_link}' target='_blank' rel='noopener noreferrer'>PDF</a></p>"""

    # Create news
    news = News(
        title=f"Résultats – {call.title}",
        content=content,
        excerpt=f"Les résultats de l'appel « {call.title} » ({call.reference_number}) sont disponibles.",
        is_published=True,
        is_featured=True,
        published_at=datetime.now(timezone.utc),
        created_by_id=current_user.id,
        call_id=call.id,
    )
    session.add(news)

    # Optionally update call status
    from ..db.models import CallStatus
    call.status = CallStatus.RESULTS_PUBLISHED
    call.results_publication_date = datetime.now(timezone.utc)

    await session.commit()

    return {
        "message": "Résultats publiés avec succès comme actualité",
        "news_id": news.id,
    }
